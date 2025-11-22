"""Endpoints de gestion des utilisateurs."""
from fastapi import APIRouter, Depends, HTTPException, status, Request, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from uuid import UUID
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

from app.db.session import get_db
from app.api.deps import get_current_user, require_admin
from app.schemas.user import (
    UserCreate,
    UserUpdate,
    UserResponse,
    UserListResponse,
    ResetPasswordRequest,
    UserImportResult,
    UserStatsResponse
)
from app.services import UserService
from app.models.user import User, UserRole

router = APIRouter()


@router.get("/import-excel/template", summary="Télécharger le template Excel")
async def download_import_template(
    current_user: User = Depends(require_admin)
):
    """
    Génère et télécharge un template Excel pré-formaté pour l'import d'utilisateurs.
    
    **Permissions**: ADMIN uniquement
    
    **Format du template** :
    - Ligne 1 : Headers en gras avec couleur de fond
    - Ligne 2 : Exemple d'utilisateur avec données fictives
    - Colonnes :
      - A: matricule (ex: U0001)
      - B: email (ex: utilisateur@beac.int)
      - C: nom (ex: Dupont)
      - D: prenom (ex: Jean)
      - E: role (ADMIN, MANAGER, USER)
      - F: password (min 10 caractères)
    
    Returns:
        Fichier Excel (.xlsx) prêt à remplir
    """
    # Créer un nouveau workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Utilisateurs"
    
    # Définir les styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="005CA9", end_color="005CA9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    example_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["matricule", "email", "nom", "prenom", "role", "password"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Exemple de données (ligne 2)
    example_data = [
        "U0001",
        "jean.dupont@beac.int",
        "Dupont",
        "Jean",
        "USER",
        "MonMotDePasse123!"
    ]
    
    for col_num, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = value
        cell.fill = example_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Ajuster la largeur des colonnes
    column_widths = {
        'A': 15,  # matricule
        'B': 30,  # email
        'C': 20,  # nom
        'D': 20,  # prenom
        'E': 12,  # role
        'F': 25   # password
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Ajouter une feuille d'instructions
    ws_instructions = wb.create_sheet("Instructions")
    
    instructions = [
        ["INSTRUCTIONS POUR L'IMPORT D'UTILISATEURS"],
        [""],
        ["1. Format du fichier :"],
        ["   - Ne pas modifier les headers de la ligne 1"],
        ["   - Commencer à remplir à partir de la ligne 3"],
        ["   - La ligne 2 contient un exemple (vous pouvez la supprimer)"],
        [""],
        ["2. Colonnes obligatoires :"],
        ["   - matricule : Identifiant unique (ex: U0001, M0042, A0003)"],
        ["   - email : Adresse email valide et unique"],
        ["   - nom : Nom de famille"],
        ["   - prenom : Prénom"],
        ["   - role : Un des trois rôles (ADMIN, MANAGER, USER)"],
        ["   - password : Mot de passe (minimum 10 caractères)"],
        [""],
        ["3. Règles de validation :"],
        ["   - Le matricule doit être unique dans le système"],
        ["   - L'email doit être unique et valide"],
        ["   - Le mot de passe doit contenir :"],
        ["     * Au moins 10 caractères"],
        ["     * Au moins une majuscule"],
        ["     * Au moins une minuscule"],
        ["     * Au moins un chiffre"],
        ["     * Au moins un caractère spécial (!@#$%^&*(),.?\":{}|<>)"],
        ["   - Le rôle doit être exactement : ADMIN, MANAGER ou USER"],
        [""],
        ["4. Rôles disponibles :"],
        ["   - ADMIN : Accès complet au système"],
        ["   - MANAGER : Gestion des documents et catégories"],
        ["   - USER : Utilisation du chatbot uniquement"],
        [""],
        ["5. Conseils :"],
        ["   - Utilisez des matricules cohérents (ex: U pour User, M pour Manager, A pour Admin)"],
        ["   - Les mots de passe temporaires devront être changés par les utilisateurs"],
        ["   - En cas d'erreur, seuls les utilisateurs en erreur seront ignorés"],
        ["   - Le rapport d'import vous indiquera les succès et les erreurs"],
        [""],
        ["6. Exemple de données :"],
        ["   Voir la ligne 2 de la feuille 'Utilisateurs' pour un exemple complet"],
    ]
    
    # Styles pour les instructions
    title_font = Font(bold=True, size=14, color="005CA9")
    section_font = Font(bold=True, size=11)
    
    for row_num, instruction in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row_num, column=1)
        cell.value = instruction[0]
        
        # Style du titre
        if row_num == 1:
            cell.font = title_font
        # Style des sections (lignes qui commencent par un chiffre suivi d'un point)
        elif instruction[0] and instruction[0][0].isdigit():
            cell.font = section_font
    
    # Ajuster la largeur de la colonne des instructions
    ws_instructions.column_dimensions['A'].width = 80
    
    # Sauvegarder le workbook dans un BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Retourner le fichier en tant que StreamingResponse
    headers = {
        'Content-Disposition': 'attachment; filename="template_import_utilisateurs.xlsx"'
    }
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


@router.get("", response_model=UserListResponse, summary="Liste des utilisateurs")
async def get_users(
    skip: int = Query(0, ge=0, description="Nombre d'éléments à sauter"),
    limit: int = Query(20, ge=1, le=100, description="Nombre d'éléments par page"),
    search: Optional[str] = Query(None, description="Recherche (matricule, nom, prénom, email)"),
    role: Optional[UserRole] = Query(None, description="Filtrer par rôle"),
    is_active: Optional[bool] = Query(None, description="Filtrer par statut actif/inactif"),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Récupère la liste paginée des utilisateurs avec filtres.
    
    **Permissions**: ADMIN uniquement
    
    Query params:
    - **skip**: Pagination - nombre d'éléments à sauter
    - **limit**: Pagination - nombre d'éléments par page (max 100)
    - **search**: Recherche dans matricule, nom, prénom, email
    - **role**: Filtrer par rôle (ADMIN, MANAGER, USER)
    - **is_active**: Filtrer par statut actif (true/false)
    
    Returns:
        Liste paginée d'utilisateurs avec métadonnées de pagination
    """
    users, total = UserService.get_users(
        db=db,
        skip=skip,
        limit=limit,
        search=search,
        role=role,
        is_active=is_active
    )
    
    # Calculer le nombre total de pages
    total_pages = math.ceil(total / limit) if total > 0 else 0
    page = (skip // limit) + 1
    
    return UserListResponse(
        users=[UserResponse.model_validate(user) for user in users],
        total=total,
        page=page,
        page_size=limit,
        total_pages=total_pages
    )


@router.post("", response_model=UserResponse, status_code=status.HTTP_201_CREATED, summary="Créer un utilisateur")
async def create_user(
    user_data: UserCreate,
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Crée un nouvel utilisateur.
    
    **Permissions**: ADMIN uniquement
    
    Body:
    - **matricule**: Matricule unique
    - **email**: Email unique
    - **nom**: Nom de famille
    - **prenom**: Prénom
    - **password**: Mot de passe (min 10 caractères)
    - **role**: Rôle (ADMIN, MANAGER, USER)
    - **is_active**: Statut actif (default: true)
    
    Returns:
        Utilisateur créé
    """
    ip_address = request.client.host if request.client else None
    
    try:
        user = UserService.create_user(
            db=db,
            user_data=user_data,
            created_by=current_user.id,
            ip_address=ip_address
        )
        
        return UserResponse.model_validate(user)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erreur lors de la création de l'utilisateur: {str(e)}"
        )


@router.get("/{user_id}", response_model=UserResponse, summary="Détails d'un utilisateur")
async def get_user(
    user_id: UUID,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Récupère les détails d'un utilisateur par son ID.
    
    **Permissions**: ADMIN uniquement
    
    Path params:
    - **user_id**: ID UUID de l'utilisateur
    
    Returns:
        Détails de l'utilisateur
    """
    user = UserService.get_user_by_id(db, user_id)
    
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Utilisateur avec l'ID {user_id} introuvable"
        )
    
    return UserResponse.model_validate(user)


@router.put("/{user_id}", response_model=UserResponse, summary="Modifier un utilisateur")
async def update_user(
    user_id: UUID,
    user_data: UserUpdate,
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Met à jour un utilisateur existant.
    
    **Permissions**: ADMIN uniquement
    
    Path params:
    - **user_id**: ID UUID de l'utilisateur
    
    Body (tous les champs optionnels):
    - **email**: Nouvel email
    - **nom**: Nouveau nom
    - **prenom**: Nouveau prénom
    - **role**: Nouveau rôle
    - **is_active**: Nouveau statut actif
    
    Returns:
        Utilisateur mis à jour
    """
    ip_address = request.client.host if request.client else None
    
    try:
        user = UserService.update_user(
            db=db,
            user_id=user_id,
            user_data=user_data,
            updated_by=current_user.id,
            ip_address=ip_address
        )
        
        return UserResponse.model_validate(user)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erreur lors de la mise à jour de l'utilisateur: {str(e)}"
        )


@router.delete("/{user_id}", status_code=status.HTTP_204_NO_CONTENT, summary="Supprimer un utilisateur")
async def delete_user(
    user_id: UUID,
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Supprime un utilisateur.
    
    **Permissions**: ADMIN uniquement
    
    **Protection**: Impossible de supprimer le dernier admin actif
    
    Path params:
    - **user_id**: ID UUID de l'utilisateur à supprimer
    
    Returns:
        204 No Content en cas de succès
    """
    ip_address = request.client.host if request.client else None
    
    try:
        UserService.delete_user(
            db=db,
            user_id=user_id,
            deleted_by=current_user.id,
            ip_address=ip_address
        )
        
        return None
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erreur lors de la suppression de l'utilisateur: {str(e)}"
        )


@router.post("/import-excel", response_model=UserImportResult, summary="Importer des utilisateurs depuis Excel")
async def import_users_from_excel(
    file: UploadFile = File(..., description="Fichier Excel (.xlsx)"),
    request: Request = None,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Importe des utilisateurs en masse depuis un fichier Excel.
    
    **Permissions**: ADMIN uniquement
    
    **Format Excel attendu** (ligne 1 = headers):
    - Colonne A: matricule
    - Colonne B: email
    - Colonne C: nom
    - Colonne D: prenom
    - Colonne E: role (ADMIN, MANAGER, USER)
    - Colonne F: password
    
    **Conseil**: Téléchargez le template via GET /users/import-excel/template
    
    File:
    - **file**: Fichier Excel (.xlsx)
    
    Returns:
        Résultat de l'import avec nombre de succès/erreurs et détails
    """
    # Vérifier l'extension du fichier
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Le fichier doit être au format Excel (.xlsx ou .xls)"
        )
    
    ip_address = request.client.host if request and request.client else None
    
    try:
        result = await UserService.import_users_from_excel(
            db=db,
            file=file,
            imported_by=current_user.id,
            ip_address=ip_address
        )
        
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erreur lors de l'import Excel: {str(e)}"
        )


@router.post("/{user_id}/reset-password", summary="Réinitialiser le mot de passe")
async def reset_user_password(
    user_id: UUID,
    password_data: ResetPasswordRequest,
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Réinitialise le mot de passe d'un utilisateur (admin only).
    
    **Permissions**: ADMIN uniquement
    
    Path params:
    - **user_id**: ID UUID de l'utilisateur
    
    Body:
    - **new_password**: Nouveau mot de passe (min 10 caractères)
    - **force_change**: Forcer le changement au prochain login (default: true)
    
    Returns:
        Message de confirmation
    """
    ip_address = request.client.host if request.client else None
    
    try:
        UserService.reset_password(
            db=db,
            user_id=user_id,
            new_password=password_data.new_password,
            reset_by=current_user.id,
            force_change=password_data.force_change,
            ip_address=ip_address
        )
        
        return {
            "message": "Mot de passe réinitialisé avec succès",
            "force_change": password_data.force_change,
            "detail": "L'utilisateur devra changer son mot de passe au prochain login" if password_data.force_change else None
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erreur lors de la réinitialisation du mot de passe: {str(e)}"
        )


@router.get("/stats/overview", response_model=UserStatsResponse, summary="Statistiques des utilisateurs")
async def get_user_stats(
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Récupère les statistiques globales des utilisateurs.
    
    **Permissions**: ADMIN uniquement
    
    Returns:
        Statistiques:
        - Total utilisateurs
        - Utilisateurs actifs/inactifs
        - Répartition par rôle
        - Connexions récentes (7 derniers jours)
    """
    try:
        stats = UserService.get_user_stats(db)
        return stats
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erreur lors de la récupération des statistiques: {str(e)}"
        )